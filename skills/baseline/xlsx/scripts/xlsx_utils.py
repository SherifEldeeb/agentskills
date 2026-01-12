#!/usr/bin/env python3
"""
XLSX Utility Functions

Common utilities for working with Excel spreadsheets.

Usage:
    from xlsx_utils import read_workbook, create_workbook, add_chart
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Any, Union

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    raise ImportError("openpyxl is required. Install with: pip install openpyxl")

try:
    import pandas as pd
except ImportError:
    raise ImportError("pandas is required. Install with: pip install pandas")


logger = logging.getLogger(__name__)


# Style presets
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

SEVERITY_COLORS = {
    'Critical': 'E74C3C',
    'High': 'E67E22',
    'Medium': 'F1C40F',
    'Low': '3498DB',
    'Info': '95A5A6'
}


def read_workbook(file_path: Path, data_only: bool = True) -> Dict[str, List[List[Any]]]:
    """
    Read all sheets from an Excel workbook.

    Args:
        file_path: Path to the Excel file
        data_only: If True, read calculated values instead of formulas

    Returns:
        Dictionary mapping sheet names to data (list of rows)
    """
    wb = load_workbook(file_path, data_only=data_only)
    data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = []

        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                sheet_data.append(list(row))

        data[sheet_name] = sheet_data

    return data


def read_sheet(file_path: Path, sheet_name: str = None, data_only: bool = True) -> List[List[Any]]:
    """
    Read a specific sheet from an Excel workbook.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of sheet to read (default: active sheet)
        data_only: If True, read calculated values instead of formulas

    Returns:
        List of rows (each row is a list of cell values)
    """
    wb = load_workbook(file_path, data_only=data_only)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    data = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            data.append(list(row))

    return data


def read_to_dataframe(file_path: Path, sheet_name: str = None) -> pd.DataFrame:
    """
    Read Excel sheet into a pandas DataFrame.

    Args:
        file_path: Path to the Excel file
        sheet_name: Name of sheet to read (default: first sheet)

    Returns:
        pandas DataFrame
    """
    return pd.read_excel(file_path, sheet_name=sheet_name)


def read_all_sheets_to_dataframes(file_path: Path) -> Dict[str, pd.DataFrame]:
    """
    Read all sheets from Excel into DataFrames.

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary mapping sheet names to DataFrames
    """
    xlsx = pd.ExcelFile(file_path)
    return {name: pd.read_excel(xlsx, sheet_name=name) for name in xlsx.sheet_names}


def create_workbook(
    data: Dict[str, List[List[Any]]],
    output_path: Path,
    format_headers: bool = True
) -> bool:
    """
    Create a new Excel workbook with multiple sheets.

    Args:
        data: Dictionary mapping sheet names to data
        output_path: Path for output file
        format_headers: Whether to format the first row as headers

    Returns:
        True if successful
    """
    try:
        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, sheet_data in data.items():
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit

            for row_idx, row in enumerate(sheet_data, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = THIN_BORDER

                    if format_headers and row_idx == 1:
                        cell.font = HEADER_FONT
                        cell.fill = HEADER_FILL
                        cell.alignment = Alignment(horizontal='center')

            # Auto-fit columns
            for column in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in column)
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

        wb.save(output_path)
        return True
    except Exception as e:
        logger.error(f"Failed to create workbook: {e}")
        return False


def dataframe_to_workbook(
    dataframes: Dict[str, pd.DataFrame],
    output_path: Path,
    format_headers: bool = True
) -> bool:
    """
    Write multiple DataFrames to an Excel workbook.

    Args:
        dataframes: Dictionary mapping sheet names to DataFrames
        output_path: Path for output file
        format_headers: Whether to format headers

    Returns:
        True if successful
    """
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for name, df in dataframes.items():
                df.to_excel(writer, sheet_name=name[:31], index=False)

        if format_headers:
            wb = load_workbook(output_path)
            for ws in wb.worksheets:
                for cell in ws[1]:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                    cell.alignment = Alignment(horizontal='center')
            wb.save(output_path)

        return True
    except Exception as e:
        logger.error(f"Failed to create workbook: {e}")
        return False


def add_sheet(
    workbook_path: Path,
    sheet_name: str,
    data: List[List[Any]],
    format_headers: bool = True
) -> bool:
    """
    Add a new sheet to an existing workbook.

    Args:
        workbook_path: Path to existing workbook
        sheet_name: Name for the new sheet
        data: Data to write
        format_headers: Whether to format headers

    Returns:
        True if successful
    """
    try:
        try:
            wb = load_workbook(workbook_path)
        except FileNotFoundError:
            wb = Workbook()
            wb.remove(wb.active)

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

        ws = wb.create_sheet(title=sheet_name[:31])

        for row_idx, row in enumerate(data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER

                if format_headers and row_idx == 1:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL

        wb.save(workbook_path)
        return True
    except Exception as e:
        logger.error(f"Failed to add sheet: {e}")
        return False


def add_chart(
    workbook_path: Path,
    sheet_name: str,
    chart_type: str,
    data_range: tuple,
    category_range: tuple,
    title: str,
    position: str = 'E2'
) -> bool:
    """
    Add a chart to a worksheet.

    Args:
        workbook_path: Path to workbook
        sheet_name: Sheet to add chart to
        chart_type: Type of chart ('bar', 'pie', 'line')
        data_range: Tuple of (min_col, min_row, max_col, max_row)
        category_range: Tuple of (min_col, min_row, max_col, max_row)
        title: Chart title
        position: Cell position for chart

    Returns:
        True if successful
    """
    try:
        wb = load_workbook(workbook_path)
        ws = wb[sheet_name]

        chart_classes = {
            'bar': BarChart,
            'pie': PieChart,
            'line': LineChart
        }

        chart_class = chart_classes.get(chart_type.lower(), BarChart)
        chart = chart_class()
        chart.title = title

        data_ref = Reference(ws, *data_range)
        cat_ref = Reference(ws, *category_range)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)

        ws.add_chart(chart, position)
        wb.save(workbook_path)
        return True
    except Exception as e:
        logger.error(f"Failed to add chart: {e}")
        return False


def add_data_validation(
    workbook_path: Path,
    sheet_name: str,
    column: str,
    options: List[str],
    start_row: int = 2,
    end_row: int = 1000
) -> bool:
    """
    Add dropdown validation to a column.

    Args:
        workbook_path: Path to workbook
        sheet_name: Sheet name
        column: Column letter (e.g., 'B')
        options: List of valid options
        start_row: First row to apply validation
        end_row: Last row to apply validation

    Returns:
        True if successful
    """
    try:
        wb = load_workbook(workbook_path)
        ws = wb[sheet_name]

        dv = DataValidation(
            type='list',
            formula1=f'"{",".join(options)}"',
            allow_blank=True,
            showErrorMessage=True
        )

        ws.add_data_validation(dv)
        dv.add(f'{column}{start_row}:{column}{end_row}')

        wb.save(workbook_path)
        return True
    except Exception as e:
        logger.error(f"Failed to add data validation: {e}")
        return False


def apply_conditional_formatting(
    workbook_path: Path,
    sheet_name: str,
    column: str,
    start_row: int = 2,
    end_row: int = 1000
) -> bool:
    """
    Apply severity-based conditional formatting to a column.

    Args:
        workbook_path: Path to workbook
        sheet_name: Sheet name
        column: Column letter
        start_row: First row
        end_row: Last row

    Returns:
        True if successful
    """
    from openpyxl.formatting.rule import CellIsRule

    try:
        wb = load_workbook(workbook_path)
        ws = wb[sheet_name]

        cell_range = f'{column}{start_row}:{column}{end_row}'

        for severity, color in SEVERITY_COLORS.items():
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            rule = CellIsRule(operator='equal', formula=[f'"{severity}"'], fill=fill)
            ws.conditional_formatting.add(cell_range, rule)

        wb.save(workbook_path)
        return True
    except Exception as e:
        logger.error(f"Failed to apply conditional formatting: {e}")
        return False


def merge_workbooks(
    workbook_paths: List[Path],
    output_path: Path,
    prefix_sheets: bool = True
) -> bool:
    """
    Merge multiple workbooks into one.

    Args:
        workbook_paths: List of workbook paths to merge
        output_path: Output file path
        prefix_sheets: Add source filename prefix to sheet names

    Returns:
        True if successful
    """
    try:
        merged_wb = Workbook()
        merged_wb.remove(merged_wb.active)

        for wb_path in workbook_paths:
            source_wb = load_workbook(wb_path)
            prefix = wb_path.stem[:10] + '_' if prefix_sheets else ''

            for sheet_name in source_wb.sheetnames:
                source_ws = source_wb[sheet_name]
                new_name = (prefix + sheet_name)[:31]

                # Ensure unique name
                counter = 1
                base_name = new_name
                while new_name in merged_wb.sheetnames:
                    new_name = f"{base_name[:28]}_{counter}"
                    counter += 1

                new_ws = merged_wb.create_sheet(title=new_name)

                for row in source_ws.iter_rows():
                    for cell in row:
                        new_ws[cell.coordinate].value = cell.value

        merged_wb.save(output_path)
        return True
    except Exception as e:
        logger.error(f"Failed to merge workbooks: {e}")
        return False


def get_workbook_info(file_path: Path) -> Dict[str, Any]:
    """
    Get information about an Excel workbook.

    Args:
        file_path: Path to the Excel file

    Returns:
        Dictionary with workbook information
    """
    wb = load_workbook(file_path, data_only=True)

    info = {
        'sheets': [],
        'total_sheets': len(wb.sheetnames)
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        info['sheets'].append({
            'name': sheet_name,
            'rows': ws.max_row,
            'columns': ws.max_column
        })

    return info


def export_to_csv(
    workbook_path: Path,
    output_dir: Path,
    sheet_name: Optional[str] = None
) -> List[Path]:
    """
    Export Excel sheets to CSV files.

    Args:
        workbook_path: Path to Excel file
        output_dir: Directory for CSV files
        sheet_name: Specific sheet to export (None for all)

    Returns:
        List of created CSV file paths
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    created_files = []

    xlsx = pd.ExcelFile(workbook_path)
    sheets = [sheet_name] if sheet_name else xlsx.sheet_names

    for name in sheets:
        df = pd.read_excel(xlsx, sheet_name=name)
        csv_path = output_dir / f"{name}.csv"
        df.to_csv(csv_path, index=False)
        created_files.append(csv_path)

    return created_files
